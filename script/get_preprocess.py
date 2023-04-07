import polars as pl
import pandas as pd
import fire
import openpyxl


def get_preprocess(START_YEAR: int = 2020
                   , END_YEAR: int = 2023
                   , PATH: str = "./datasets/"):
	dfs = []
	start_year = START_YEAR
	end_year = END_YEAR

	for year in range(start_year, end_year + 1):
		file_path = PATH + f"internal_{year}.xlsx"
		workbook = openpyxl.load_workbook(file_path, read_only=True)
		num_sheets = len(workbook.sheetnames)

		for sheet_no in range(1, num_sheets + 1):
			for suffix in ["_2", "_1"] if year == start_year else ["_2"]:

				cleaned_columns_list = [f'province{suffix}'
					, 'ratio_stay_2', 'no_tourist_1', 'no_tourist_change'
					, 'no_stay_2', 'no_stay_1', 'no_stay_change'
					, 'no_tourist_2', 'no_tourist_1', 'no_tourist_change'
					, 'no_tourist_thai_2', 'no_tourist_thai_1', 'no_tourist_thai_change'
					, 'no_tourist_foreign_2', 'no_tourist_foreign_1', 'no_tourist_foreign_change'
					, 'profit_all_2', 'profit_all_1', 'profit_all_change'
					, 'profit_tourist_thai_2', 'profit_tourist_thai_1', 'profit_tourist_thai_change'
					, 'profit_tourist_foreign_2', 'profit_tourist_foreign_1', 'profit_tourist_foreign_change']

				raw_excel = (
					pl.read_excel(
						file_path,
						sheet_id=sheet_no,
						xlsx2csv_options={"skip_empty_lines": True},
					)
					.to_pandas()
					.dropna(axis=1, how="all")
				)

				raw_excel.columns = cleaned_columns_list
				raw_excel = (
					raw_excel.drop(raw_excel[raw_excel.iloc[:, 0] == "จังหวัด"].index)
					.dropna(subset=[raw_excel.columns[0]])
					.replace("(R1)", "")
					.replace("*", "")
					.loc[lambda x: x[f"province{suffix}"].str.len() < 20]
					.loc[lambda x: ~x[f"province{suffix}"].str.contains("ภาค")]
					.loc[lambda x: ~x[f"province{suffix}"].str.contains("รวม")]
					.filter(regex=suffix)
				)

				raw_excel.columns = raw_excel.columns.str.replace(suffix, "")
				raw_excel = raw_excel.astype(
					{col: "float" for col in raw_excel.columns if col != "province"}
				)
				raw_excel["time"] = workbook.sheetnames[sheet_no - 1]
				raw_excel = raw_excel.assign(
					month=raw_excel["time"].str.extract("(\D+)"),
					year=raw_excel["time"].str.extract("(\d+)"),
				).drop(columns=["time"])
				raw_excel["year"] = raw_excel["year"].astype(int)

				if suffix == "_2":
					raw_excel["year"] = year
				else:
					raw_excel["year"] = year - 1
				dfs.append(raw_excel)

	combined_df = pd.concat(dfs, ignore_index=True)
	combined_df['province'] = (combined_df['province']
	                           .str.replace(r'(R1)', "", regex=True)
	                           .str.replace(r'(\(|\))', "", regex=True)
	                           .str.replace(r'R', "", regex=True)
	                           .str.replace(r'*', "", regex=True)
	                           .str.replace(r'ประจวบศีรีขันธ์', "ประจวบคีรีขันธ์", regex=True)
	                           .str.replace(r'ศรีษะเกษ', "ศรีสะเกษ", regex=True)
	                           .str.replace(r'จ.นคร', "นคร", regex=True)
	                           .str.replace(r'^\s+', "", regex=True))

	month_index = {
		"ม.ค.": 1, "ก.พ.": 2, "มี.ค.": 3, "เม.ย.": 4, "พ.ค.": 5, "มิ.ย.": 6, "ก.ค.": 7, "ส.ค.": 8, "ก.ย.": 9,
		"ต.ค.": 10, "พ.ย.": 11, "ธ.ค.": 12
	}
	combined_df['month_no'] = combined_df['month'].map(month_index)
	combined_df = combined_df.sort_values(["month_no", "year"])
	combined_df.to_csv("./datasets/combined.csv", index=False)


if __name__ == '__main__':
	fire.Fire(get_preprocess)
